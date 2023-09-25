from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
total=0
max_row = ws.max_row
min_row =2
for now in range(2,max_row+1):
    hours = (ws ['B'+str(now)].value)
    rate = (ws ['C'+str(now)].value)
    if 'a' not in str(rate) and 'a' not in str(hours):
        salary = int(rate) * int(hours)
        if salary > (3000): 
            total+=1
    

print(total)
wb.close()
